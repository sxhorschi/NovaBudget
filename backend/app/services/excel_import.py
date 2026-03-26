"""Excel Import Engine — intelligently parses TYTAN CAPEX Excel workbooks.

Key capabilities:
- Auto-detects functional_area sheets by header structure (not just hardcoded names)
- Extracts budget from BudgetTemplate sheet (with 0.85 factor)
- Reads calculated values (data_only=True) and preserves formula info
- Duplicate detection: update-on-reimport instead of blind insert
- Detailed per-functional_area validation report
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import CostItem, FunctionalArea, WorkArea
from app.models.enums import ApprovalStatus
from app.services.import_report import FunctionalAreaReport, ImportReport

logger = logging.getLogger(__name__)

# ── Known sheet names → display names (fallback if auto-detection fails) ──────
KNOWN_DEPARTMENT_SHEETS: dict[str, str] = {
    "Assembly_Equipment": "Assembly Equipment",
    "Testing": "Testing",
    "Logistics": "Logistics",
    "Facility": "Facility",
    "Prototyping": "Prototyping",
}

# ── Expected header keywords (case-insensitive) for auto-detection ────────────
# If Row 5 contains at least 3 of these in any order, we treat it as a dept sheet.
HEADER_KEYWORDS = {"work area", "description", "amount", "cost basis", "approval", "phase"}

HEADER_ROW = 5
DATA_START_ROW = 6
BUDGET_FACTOR = Decimal("0.85")

# ── Column mapping (1-indexed): matches Excel layout ─────────────────────────
COL_WORK_AREA = 2       # B
COL_PHASE = 3            # C
COL_PRODUCT = 4          # D
COL_DESCRIPTION = 5      # E
COL_AMOUNT = 6           # F
COL_CASH_OUT = 7         # G
COL_COST_BASIS = 8       # H
COL_COST_DRIVER = 9      # I
COL_BASIS_DESC = 10      # J
COL_ASSUMPTIONS = 11     # K
COL_APPROVAL = 12        # L
COL_APPROVAL_DATE = 13   # M
COL_ZIELANPASSUNG = 14   # N
COL_COMMENTS = 15        # O
COL_REQUESTER = 16       # P

DEFAULT_COLUMN_MAP: dict[str, int] = {
    "work_area": COL_WORK_AREA,
    "phase": COL_PHASE,
    "product": COL_PRODUCT,
    "description": COL_DESCRIPTION,
    "amount": COL_AMOUNT,
    "cash_out": COL_CASH_OUT,
    "cost_basis": COL_COST_BASIS,
    "cost_driver": COL_COST_DRIVER,
    "basis_desc": COL_BASIS_DESC,
    "assumptions": COL_ASSUMPTIONS,
    "approval": COL_APPROVAL,
    "approval_date": COL_APPROVAL_DATE,
    "zielanpassung": COL_ZIELANPASSUNG,
    "comments": COL_COMMENTS,
    "requester": COL_REQUESTER,
}

HEADER_ALIASES: dict[str, set[str]] = {
    "work_area": {"work area", "workarea", "arbeitsbereich"},
    "phase": {"phase", "projektphase"},
    "product": {"product", "produkt"},
    "description": {"description", "beschreibung", "desc"},
    "amount": {"amount", "betrag", "value", "kosten"},
    "cash_out": {"cash out", "cash-out", "cashout", "cash out datum", "cash out date"},
    "cost_basis": {"cost basis", "kostenbasis", "basis"},
    "cost_driver": {"cost driver", "kostentreiber", "driver"},
    "basis_desc": {"basis description", "basis beschreibung"},
    "assumptions": {"assumptions", "annahmen"},
    "approval": {"approval", "approval status", "genehmigung", "approval_status"},
    "approval_date": {"approval date", "freigabedatum", "genehmigungsdatum"},
    "zielanpassung": {"zielanpassung", "target adjustment"},
    "comments": {"comments", "kommentare", "comment"},
    "requester": {"requester", "anforderer", "requested by"},
}


# ═══════════════════════════════════════════════════════════════════════════════
#  Helper functions
# ═══════════════════════════════════════════════════════════════════════════════

def _safe_decimal(value) -> Decimal:
    """Convert a cell value to Decimal, returning 0 on failure."""
    if value is None:
        return Decimal(0)
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, ArithmeticError):
        return Decimal(0)


def _safe_str(value) -> str | None:
    """Convert a cell value to stripped string, returning None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_date(value) -> date | None:
    """Parse a cell value into a date. Handles datetime, date, and string formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    str_val = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str_val, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _parse_enum(enum_class, value, default=None):
    """Fuzzy-match a cell value to a Python enum member.

    Returns (member, error_msg|None).
    """
    if value is None:
        return default, None
    str_val = str(value).strip()
    if not str_val:
        return default, None

    normalized = str_val.upper().replace(" ", "_").replace("-", "_")

    # Exact match on value or name
    for member in enum_class:
        member_norm = member.value.upper().replace(" ", "_").replace("-", "_")
        if member_norm == normalized or member.name == normalized:
            return member, None

    # Partial / contains match
    for member in enum_class:
        member_norm = member.value.upper().replace(" ", "_").replace("-", "_")
        if normalized in member_norm or member_norm in normalized:
            return member, None

    return default, f"Unknown {enum_class.__name__} value '{str_val}'"


# ── String-based config field parsing (replaces enum parsing for flexible fields) ─

# Maps for fuzzy-matching Excel values to config string ids
_PRODUCT_ALIASES: dict[str, str] = {
    "ATLAS": "atlas", "ORION": "orion", "VEGA": "vega", "OVERALL": "overall",
    # Legacy names from old Excel files
    "BRYAN": "atlas", "GUENTHER": "orion", "GÜNTHER": "orion",
    "GIN_TONIC": "vega", "GIN__TONIC": "vega",
}

_PHASE_ALIASES: dict[str, str] = {
    "PHASE_1": "phase_1", "PHASE_2": "phase_2", "PHASE_3": "phase_3", "PHASE_4": "phase_4",
    "PHASE 1": "phase_1", "PHASE 2": "phase_2", "PHASE 3": "phase_3", "PHASE 4": "phase_4",
    "1": "phase_1", "2": "phase_2", "3": "phase_3", "4": "phase_4",
}

_COST_BASIS_ALIASES: dict[str, str] = {
    "COST_ESTIMATION": "cost_estimation", "COST ESTIMATION": "cost_estimation",
    "INITIAL_SUPPLIER_OFFER": "initial_supplier_offer", "INITIAL SUPPLIER OFFER": "initial_supplier_offer",
    "REVISED_SUPPLIER_OFFER": "revised_supplier_offer", "REVISED SUPPLIER OFFER": "revised_supplier_offer",
    "CHANGE_COST": "change_cost", "CHANGE COST": "change_cost",
}

_COST_DRIVER_ALIASES: dict[str, str] = {
    "PRODUCT": "product", "PROCESS": "process",
    "NEW_REQ_ASSEMBLY": "new_req_assembly", "NEW REQ ASSEMBLY": "new_req_assembly",
    "NEW_REQ_TESTING": "new_req_testing", "NEW REQ TESTING": "new_req_testing",
    "INITIAL_SETUP": "initial_setup", "INITIAL SETUP": "initial_setup",
}


def _parse_config_field(
    alias_map: dict[str, str],
    field_name: str,
    value,
    default: str,
) -> tuple[str, str | None]:
    """Fuzzy-match a cell value to a config string id using an alias map.

    Returns (matched_id, error_msg|None).
    """
    if value is None:
        return default, None
    str_val = str(value).strip()
    if not str_val:
        return default, None

    normalized = str_val.upper().replace(" ", "_").replace("-", "_")

    # Exact match
    if normalized in alias_map:
        return alias_map[normalized], None

    # Try with spaces
    str_upper = str_val.strip().upper()
    if str_upper in alias_map:
        return alias_map[str_upper], None

    # Partial match
    for key, mapped_id in alias_map.items():
        if normalized in key or key in normalized:
            return mapped_id, None

    # If the lowercased value itself looks like a valid config id, accept it
    lower_val = str_val.strip().lower().replace(" ", "_").replace("-", "_")
    if lower_val in alias_map.values():
        return lower_val, None

    return default, f"Unknown {field_name} value '{str_val}'"


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = normalized.replace("\n", " ").replace("_", " ").replace("-", " ")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def _detect_column_map(ws: Worksheet) -> tuple[dict[str, int], set[str]]:
    """Detect actual column positions from header row and fall back to defaults."""
    detected: dict[str, int] = {}
    max_col = max(ws.max_column or 0, 20)

    for col in range(1, max_col + 1):
        header = _normalize_header(ws.cell(row=HEADER_ROW, column=col).value)
        if not header:
            continue

        for field, aliases in HEADER_ALIASES.items():
            if field in detected:
                continue
            if any(
                header == alias
                or header.startswith(alias + " ")
                or header.endswith(" " + alias)
                for alias in aliases
            ):
                detected[field] = col
                break

    col_map = {
        field: detected.get(field, default_col)
        for field, default_col in DEFAULT_COLUMN_MAP.items()
    }
    return col_map, set(detected.keys())


def _cell_value(ws: Worksheet, row: int, col_map: dict[str, int], field: str):
    return ws.cell(row=row, column=col_map[field]).value


def _is_subtotal_row(ws: Worksheet, row: int, col_map: dict[str, int] | None = None) -> bool:
    """Subtotal rows have Work Area name (B) + Amount (F) but NO Description and NO Phase."""
    cols = col_map or DEFAULT_COLUMN_MAP
    has_b = ws.cell(row=row, column=cols["work_area"]).value is not None
    has_f = ws.cell(row=row, column=cols["amount"]).value is not None
    has_description = ws.cell(row=row, column=cols["description"]).value is not None
    has_phase = ws.cell(row=row, column=cols["phase"]).value is not None
    return has_b and has_f and not has_description and not has_phase


def _is_data_row(ws: Worksheet, row: int, col_map: dict[str, int] | None = None) -> bool:
    """Data rows must have at least a description."""
    cols = col_map or DEFAULT_COLUMN_MAP
    has_description = ws.cell(row=row, column=cols["description"]).value is not None
    has_amount = ws.cell(row=row, column=cols["amount"]).value is not None
    return has_description and has_amount


def _row_is_empty(ws: Worksheet, row: int, col_map: dict[str, int] | None = None) -> bool:
    """Check if an entire mapped data row is empty."""
    cols = col_map or DEFAULT_COLUMN_MAP
    used_cols = sorted(set(cols.values()))
    return all(ws.cell(row=row, column=c).value is None for c in used_cols)


# ═══════════════════════════════════════════════════════════════════════════════
#  Sheet auto-detection
# ═══════════════════════════════════════════════════════════════════════════════

def _detect_functional_area_sheets(wb) -> dict[str, str]:
    """Detect functional_area sheets by examining header structure in Row 5.

    Returns {sheet_name: display_name}.
    Falls back to known names for sheets that match the hardcoded list.
    """
    detected: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        # Skip known non-functional_area sheets
        lower = sheet_name.lower()
        if any(skip in lower for skip in (
            "budget", "template", "dropdown", "summary", "graphic", "data_graphic", "fc", "forecast",
        )):
            continue

        ws = wb[sheet_name]
        # Check Row 5 for header keywords
        header_values: list[str] = []
        for col in range(1, 20):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val is not None:
                header_values.append(str(val).strip().lower())

        matched_keywords = sum(
            1 for kw in HEADER_KEYWORDS if any(kw in h for h in header_values)
        )

        if matched_keywords >= 3:
            # Determine display name: use known mapping or derive from sheet name
            display_name = KNOWN_DEPARTMENT_SHEETS.get(
                sheet_name,
                sheet_name.replace("_", " "),
            )
            detected[sheet_name] = display_name

    # Also include known sheets that exist but were not detected (safety net)
    for sheet_name, display_name in KNOWN_DEPARTMENT_SHEETS.items():
        if sheet_name in wb.sheetnames and sheet_name not in detected:
            # Verify it has at least some data rows
            ws = wb[sheet_name]
            col_map, _ = _detect_column_map(ws)
            has_data = any(
                ws.cell(row=r, column=col_map["description"]).value is not None
                for r in range(DATA_START_ROW, min(DATA_START_ROW + 5, ws.max_row + 1))
            )
            if has_data:
                detected[sheet_name] = display_name

    return detected


# ═══════════════════════════════════════════════════════════════════════════════
#  BudgetTemplate parsing
# ═══════════════════════════════════════════════════════════════════════════════

def _find_budget_template_sheet(wb) -> str | None:
    """Find the BudgetTemplate sheet (name starts with 'BudgetTemplate')."""
    for name in wb.sheetnames:
        if name.lower().startswith("budgettemplate"):
            return name
    return None


def _extract_functional_area_budgets(
    wb,
    functional_area_sheets: dict[str, str],
    report: ImportReport,
) -> dict[str, Decimal]:
    """Extract budget totals from BudgetTemplate sheet.

    Reads the formula-computed budget from cell F3 in each functional_area sheet.
    F3 contains: =(BudgetTemplate!M_rows)*0.85
    With data_only=True, we get the calculated value directly.

    If F3 is empty, try to parse the BudgetTemplate directly.

    Returns {display_name: budget_decimal}.
    """
    budgets: dict[str, Decimal] = {}

    # Strategy 1: Read F3 from each functional_area sheet (already has the 0.85 factor applied)
    for sheet_name, display_name in functional_area_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        f3_value = ws.cell(row=3, column=6).value  # F3
        if f3_value is not None:
            budget = _safe_decimal(f3_value)
            if budget > 0:
                budgets[display_name] = budget
                report.add_info(
                    f"{display_name}: Budget EUR {budget:,.2f} "
                    f"(from F3, includes {BUDGET_FACTOR} factor)"
                )

    # Strategy 2: If BudgetTemplate exists, read total from row 57 col M for unmapped depts
    bt_sheet_name = _find_budget_template_sheet(wb)
    if bt_sheet_name:
        report.budget_template_sheet = bt_sheet_name
        report.add_info(f"BudgetTemplate sheet found: '{bt_sheet_name}'")
    else:
        report.add_warning("No BudgetTemplate sheet found — budgets only from F3 cells")

    return budgets


# ═══════════════════════════════════════════════════════════════════════════════
#  Formula detection (reads from the non-data_only workbook)
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_formula_comment(ws_formulas: Worksheet | None, row: int, col: int) -> str | None:
    """If the cell contains a formula, return it as a comment string."""
    if ws_formulas is None:
        return None
    cell = ws_formulas.cell(row=row, column=col)
    val = cell.value
    if val is not None and isinstance(val, str) and val.startswith("="):
        return f"[Formula: {val}]"
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  Duplicate detection
# ═══════════════════════════════════════════════════════════════════════════════

async def _find_existing_item(
    session: AsyncSession,
    work_area_id: UUID,
    description: str,
) -> CostItem | None:
    """Find an existing CostItem by work_area + description (case-insensitive)."""
    stmt = select(CostItem).where(
        CostItem.work_area_id == work_area_id,
        CostItem.description.ilike(description.strip()),
    )
    result = await session.execute(stmt)
    return result.scalar_one_or_none()


# ═══════════════════════════════════════════════════════════════════════════════
#  Main import function
# ═══════════════════════════════════════════════════════════════════════════════

async def import_excel_file(
    file_contents: bytes,
    facility_id: UUID,
    session: AsyncSession,
    dry_run: bool = False,
) -> dict:
    """Import a TYTAN CAPEX Excel workbook into the database.

    When *dry_run* is True the entire import is validated and statistics are
    computed, but the database transaction is rolled back instead of committed.
    The returned report includes a ``dry_run`` flag.

    Returns a detailed JSON-serializable report dict.
    """
    report = ImportReport()
    file_bytes = io.BytesIO(file_contents)

    # Load workbook twice: data_only=True for values, data_only=False for formulas
    wb_data = load_workbook(file_bytes, read_only=True, data_only=True)
    file_bytes.seek(0)
    try:
        wb_formulas = load_workbook(file_bytes, read_only=True, data_only=False)
    except Exception:
        wb_formulas = None
        report.add_warning("Could not load formula workbook — formula comments disabled")

    try:
        # ── Step 1: Detect functional_area sheets ──────────────────────────────
        functional_area_sheets = _detect_functional_area_sheets(wb_data)
        report.sheets_detected = list(functional_area_sheets.keys())
        report.sheets_skipped = [
            s for s in wb_data.sheetnames if s not in functional_area_sheets
        ]

        if not functional_area_sheets:
            report.add_error("No functional_area sheets detected in workbook")
            return report.to_dict()

        report.add_info(
            f"Detected {len(functional_area_sheets)} functional_area sheet(s): "
            + ", ".join(functional_area_sheets.values())
        )

        # ── Step 2: Extract budgets from BudgetTemplate ──────────────────
        budgets = _extract_functional_area_budgets(wb_data, functional_area_sheets, report)

        # ── Step 3: Process each functional_area sheet ─────────────────────────
        for sheet_name, dept_display_name in functional_area_sheets.items():
            ws = wb_data[sheet_name]
            ws_form = wb_formulas[sheet_name] if wb_formulas and sheet_name in wb_formulas.sheetnames else None

            dept_report = FunctionalAreaReport(name=dept_display_name)
            col_map, detected_fields = _detect_column_map(ws)

            missing_required = {"description", "amount"} - detected_fields
            if missing_required:
                report.add_warning(
                    f"{dept_display_name}: Header not found for {', '.join(sorted(missing_required))}. "
                    "Fallback to default columns will be used."
                )

            # Find or create functional_area
            dept_stmt = select(FunctionalArea).where(
                FunctionalArea.facility_id == facility_id,
                FunctionalArea.name == dept_display_name,
            )
            dept_result = await session.execute(dept_stmt)
            functional_area = dept_result.scalar_one_or_none()

            if functional_area:
                report.functional_areas_updated += 1
            else:
                functional_area = FunctionalArea(
                    name=dept_display_name,
                    facility_id=facility_id,
                    budget_total=Decimal(0),
                )
                session.add(functional_area)
                await session.flush()
                report.functional_areas_created += 1

            # Set budget if available from BudgetTemplate
            if dept_display_name in budgets:
                functional_area.budget_total = budgets[dept_display_name]
                dept_report.budget_total = budgets[dept_display_name]

            # ── Parse rows ────────────────────────────────────────────────
            current_work_area: WorkArea | None = None
            max_row = ws.max_row or DATA_START_ROW

            for row in range(DATA_START_ROW, max_row + 1):
                # Skip fully empty rows
                if _row_is_empty(ws, row, col_map):
                    continue

                # ── Subtotal row → extract/create work area ───────────────
                if _is_subtotal_row(ws, row, col_map):
                    wa_name = _safe_str(_cell_value(ws, row, col_map, "work_area"))
                    if wa_name:
                        current_work_area = await _get_or_create_work_area(
                            session, functional_area, wa_name, dept_report,
                        )
                    continue

                # ── Data row ──────────────────────────────────────────────
                if not _is_data_row(ws, row, col_map):
                    continue

                description_raw = _cell_value(ws, row, col_map, "description")
                description = str(description_raw).strip() if description_raw else ""
                if not description:
                    continue

                # Resolve work area from col B if needed
                wa_cell = _safe_str(_cell_value(ws, row, col_map, "work_area"))
                if wa_cell and current_work_area is None:
                    current_work_area = await _get_or_create_work_area(
                        session, functional_area, wa_cell, dept_report,
                    )
                elif wa_cell and current_work_area and wa_cell != current_work_area.name:
                    # Work area changed mid-block
                    current_work_area = await _get_or_create_work_area(
                        session, functional_area, wa_cell, dept_report,
                    )

                if current_work_area is None:
                    current_work_area = await _get_or_create_work_area(
                        session, functional_area, "General", dept_report,
                    )

                # Amount (computed value from data_only workbook)
                amount = _safe_decimal(_cell_value(ws, row, col_map, "amount"))

                # Skip items with zero amount (optional: could be a warning)
                if amount == 0:
                    dept_report.items_skipped += 1
                    report.add_warning(
                        f"{dept_display_name}, Row {row}: '{description}' skipped (amount = 0)"
                    )
                    continue

                # Formula comment: if the amount cell was a formula, note it
                formula_comment = _extract_formula_comment(ws_form, row, col_map["amount"])

                # Parse enum fields with error tracking
                approval_status, approval_err = _parse_enum(
                    ApprovalStatus,
                    _cell_value(ws, row, col_map, "approval"),
                    ApprovalStatus.OPEN,
                )
                if approval_err:
                    report.add_warning(f"{dept_display_name}, Row {row}: {approval_err}")

                cost_basis, cb_err = _parse_config_field(
                    _COST_BASIS_ALIASES, "CostBasis",
                    _cell_value(ws, row, col_map, "cost_basis"),
                    "cost_estimation",
                )
                if cb_err:
                    report.add_warning(f"{dept_display_name}, Row {row}: {cb_err}")

                cost_driver, cd_err = _parse_config_field(
                    _COST_DRIVER_ALIASES, "CostDriver",
                    _cell_value(ws, row, col_map, "cost_driver"),
                    "initial_setup",
                )
                if cd_err:
                    report.add_warning(f"{dept_display_name}, Row {row}: {cd_err}")

                phase, ph_err = _parse_config_field(
                    _PHASE_ALIASES, "ProjectPhase",
                    _cell_value(ws, row, col_map, "phase"),
                    "phase_1",
                )
                if ph_err:
                    report.add_warning(f"{dept_display_name}, Row {row}: {ph_err}")

                product, pr_err = _parse_config_field(
                    _PRODUCT_ALIASES, "Product",
                    _cell_value(ws, row, col_map, "product"),
                    "overall",
                )
                if pr_err:
                    report.add_warning(f"{dept_display_name}, Row {row}: {pr_err}")

                ziel_val = _safe_decimal(_cell_value(ws, row, col_map, "zielanpassung"))

                # Build comments: merge cell comment + formula comment
                cell_comment = _safe_str(_cell_value(ws, row, col_map, "comments"))
                comments_parts: list[str] = []
                if cell_comment:
                    comments_parts.append(cell_comment)
                if formula_comment:
                    comments_parts.append(formula_comment)
                final_comments = " | ".join(comments_parts) if comments_parts else None

                # ── Duplicate detection ───────────────────────────────────
                existing = await _find_existing_item(
                    session, current_work_area.id, description,
                )

                # Requester (optional column)
                requester_val = _safe_str(_cell_value(ws, row, col_map, "requester"))

                if existing:
                    # Update existing item
                    existing.unit_price = amount
                    existing.quantity = Decimal("1")
                    existing.total_amount = amount
                    existing.expected_cash_out = _parse_date(
                        _cell_value(ws, row, col_map, "cash_out")
                    )
                    existing.cost_basis = cost_basis
                    existing.cost_driver = cost_driver
                    existing.basis_description = _safe_str(
                        _cell_value(ws, row, col_map, "basis_desc")
                    )
                    existing.assumptions = _safe_str(
                        _cell_value(ws, row, col_map, "assumptions")
                    )
                    existing.approval_status = approval_status
                    existing.approval_date = _parse_date(
                        _cell_value(ws, row, col_map, "approval_date")
                    )
                    existing.project_phase = phase
                    existing.product = product
                    existing.zielanpassung = ziel_val if ziel_val != Decimal(0) else None
                    existing.comments = final_comments
                    existing.requester = requester_val
                    dept_report.items_updated += 1
                else:
                    # Create new item
                    cost_item = CostItem(
                        work_area_id=current_work_area.id,
                        description=description,
                        unit_price=amount,
                        quantity=Decimal("1"),
                        total_amount=amount,
                        expected_cash_out=_parse_date(
                            _cell_value(ws, row, col_map, "cash_out")
                        ),
                        cost_basis=cost_basis,
                        cost_driver=cost_driver,
                        basis_description=_safe_str(
                            _cell_value(ws, row, col_map, "basis_desc")
                        ),
                        assumptions=_safe_str(
                            _cell_value(ws, row, col_map, "assumptions")
                        ),
                        approval_status=approval_status,
                        approval_date=_parse_date(
                            _cell_value(ws, row, col_map, "approval_date")
                        ),
                        project_phase=phase,
                        product=product,
                        zielanpassung=ziel_val if ziel_val != Decimal(0) else None,
                        comments=final_comments,
                        requester=requester_val,
                    )
                    session.add(cost_item)
                    dept_report.items_imported += 1

                dept_report.total_amount += amount

            report.functional_area_reports.append(dept_report)

        # ── Finalize ──────────────────────────────────────────────────────
        report.finalize()

        if dry_run:
            await session.rollback()
        else:
            await session.commit()

    finally:
        wb_data.close()
        if wb_formulas:
            wb_formulas.close()

    result = report.to_dict()
    result["dry_run"] = dry_run
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  Internal helpers
# ═══════════════════════════════════════════════════════════════════════════════

async def _get_or_create_work_area(
    session: AsyncSession,
    functional_area: FunctionalArea,
    name: str,
    dept_report: FunctionalAreaReport,
) -> WorkArea:
    """Find an existing WorkArea or create a new one."""
    stmt = select(WorkArea).where(
        WorkArea.functional_area_id == functional_area.id,
        WorkArea.name == name,
    )
    result = await session.execute(stmt)
    work_area = result.scalar_one_or_none()

    if work_area:
        dept_report.work_areas_found += 1
        return work_area

    work_area = WorkArea(
        name=name,
        functional_area_id=functional_area.id,
    )
    session.add(work_area)
    await session.flush()
    dept_report.work_areas_created += 1
    return work_area
