"""Excel export services for standard, finance, and steering committee reports."""

from __future__ import annotations

import io
import re
from datetime import date, timedelta
from decimal import Decimal
from typing import Sequence
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import CostItem, FunctionalArea, Facility, WorkArea
from app.models.enums import ApprovalStatus
from app.services.aggregation import EXCLUDED_STATUSES

# ---------------------------------------------------------------------------
# Filename helpers
# ---------------------------------------------------------------------------


def _sanitize_filename(name: str) -> str:
    """Sanitize a facility name for use in filenames.

    Replaces spaces and special characters with underscores, strips leading/trailing
    whitespace, and limits length.
    """
    sanitized = re.sub(r"[^\w\-]", "_", name.strip())
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    return sanitized[:60] if sanitized else "facility"


async def _load_facility(session: AsyncSession, facility_id: UUID) -> Facility | None:
    """Load a facility by ID."""
    stmt = select(Facility).where(Facility.id == facility_id)
    result = await session.execute(stmt)
    return result.scalar_one_or_none()


def _build_export_filename(facility_name: str, export_type: str) -> str:
    """Build a standardized export filename: TYTAN_{facility}_{type}_{YYYY-MM}.xlsx"""
    safe_name = _sanitize_filename(facility_name)
    month_str = date.today().strftime("%Y-%m")
    return f"TYTAN_{safe_name}_{export_type}_{month_str}.xlsx"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
_SUBTOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
_META_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
_NUM_FMT = '#,##0.00'
_BORDER_THIN = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
_BORDER_MEDIUM = Border(
    bottom=Side(style="medium", color="4472C4"),
)

# The configurable budget reserve factor (0.85 = 15% reserve)
DEFAULT_BUDGET_FACTOR = Decimal("0.85")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _enum_display(val) -> str:
    """Convert an enum value to a human-readable string."""
    if val is None:
        return ""
    s = val.value if hasattr(val, "value") else str(val)
    return s.replace("_", " ").title()


def _date_str(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%Y-%m-%d")


def _month_key(d: date) -> str:
    """Return 'YYYY-MM' string for a date."""
    return d.strftime("%Y-%m")


async def _load_functional_areas(
    session: AsyncSession,
    facility_id: UUID,
    functional_area_ids: list[UUID] | None = None,
) -> list[FunctionalArea]:
    """Load functional_areas with nested work_areas -> cost_items and change_costs."""
    stmt = (
        select(FunctionalArea)
        .where(FunctionalArea.facility_id == facility_id)
        .options(
            selectinload(FunctionalArea.work_areas).selectinload(WorkArea.cost_items),
            selectinload(FunctionalArea.change_costs),
        )
        .order_by(FunctionalArea.name)
    )
    if functional_area_ids:
        stmt = stmt.where(FunctionalArea.id.in_(functional_area_ids))
    result = await session.execute(stmt)
    return list(result.scalars().all())


def _fa_effective_budget(fa: FunctionalArea) -> Decimal:
    """Compute effective budget = budget_total + sum(adjustments).

    Matches the canonical definition from aggregation.py.
    """
    base = fa.budget_total or Decimal(0)
    adj = sum((a.amount or Decimal(0)) for a in fa.change_costs if a.budget_relevant)
    return base + adj


def _filter_items(
    items: Sequence[CostItem],
    phase: str | None = None,
) -> list[CostItem]:
    """Apply optional phase filter to items."""
    filtered = list(items)
    if phase:
        filtered = [i for i in filtered if i.project_phase == phase]
    return filtered


# ---------------------------------------------------------------------------
# 1. Standard Export
# ---------------------------------------------------------------------------


async def generate_standard_export(
    session: AsyncSession,
    facility_id: UUID,
    functional_area_ids: list[UUID] | None = None,
    phase: str | None = None,
) -> tuple[bytes, str]:
    """Generate a standard functional area-based Excel export.

    Returns a tuple of (xlsx_bytes, suggested_filename).
    """
    facility = await _load_facility(session, facility_id)
    facility_name = facility.name if facility else "Unknown"
    functional_areas = await _load_functional_areas(session, facility_id, functional_area_ids)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for fa in functional_areas:
        ws = wb.create_sheet(title=fa.name[:31])  # Excel sheet name max 31 chars

        # --- Header area ---
        ws.merge_cells("A1:P1")
        ws["A1"] = fa.name
        ws["A1"].font = _TITLE_FONT

        budget = _fa_effective_budget(fa)
        forecast = sum(
            item.total_amount
            for wa in fa.work_areas
            for item in _filter_items(wa.cost_items, phase)
            if item.approval_status not in EXCLUDED_STATUSES
        )
        remaining = budget - forecast

        ws["A2"] = "Budget:"
        ws["A2"].font = _META_FONT
        ws["B2"] = float(budget)
        ws["B2"].number_format = _NUM_FMT

        ws["D2"] = "Forecast:"
        ws["D2"].font = _META_FONT
        ws["E2"] = float(forecast)
        ws["E2"].number_format = _NUM_FMT

        ws["G2"] = "Remaining:"
        ws["G2"].font = _META_FONT
        ws["H2"] = float(remaining)
        ws["H2"].number_format = _NUM_FMT

        # --- Column headers (row 4) ---
        headers = [
            "Work Area", "Phase", "Product", "Description", "Amount",
            "Cash Out", "Cost Basis", "Cost Driver", "Basis Description",
            "Assumptions", "Approval Status", "Approval Date",
            "Comments", "Requester",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        row = 5
        for wa in fa.work_areas:
            items = _filter_items(wa.cost_items, phase)
            if not items:
                continue

            # Work Area Subtotal row (active items only, matching forecast definition)
            wa_total = sum(
                i.total_amount for i in items
                if i.approval_status not in EXCLUDED_STATUSES
            )
            ws.cell(row=row, column=1, value=wa.name).font = _SUBTOTAL_FONT
            for c in range(1, 15):
                ws.cell(row=row, column=c).fill = _SUBTOTAL_FILL
                ws.cell(row=row, column=c).border = _BORDER_MEDIUM
            amt_cell = ws.cell(row=row, column=5, value=float(wa_total))
            amt_cell.number_format = _NUM_FMT
            amt_cell.font = _SUBTOTAL_FONT
            row += 1

            # Individual items
            for item in items:
                ws.cell(row=row, column=1, value=wa.name)
                ws.cell(row=row, column=2, value=_enum_display(item.project_phase))
                ws.cell(row=row, column=3, value=_enum_display(item.product))
                ws.cell(row=row, column=4, value=item.description)
                amt = ws.cell(row=row, column=5, value=float(item.total_amount))
                amt.number_format = _NUM_FMT
                ws.cell(row=row, column=6, value=_date_str(item.expected_cash_out))
                ws.cell(row=row, column=7, value=_enum_display(item.cost_basis))
                ws.cell(row=row, column=8, value=_enum_display(item.cost_driver))
                ws.cell(row=row, column=9, value=item.basis_description or "")
                ws.cell(row=row, column=10, value=item.assumptions or "")
                ws.cell(row=row, column=11, value=_enum_display(item.approval_status))
                ws.cell(row=row, column=12, value=_date_str(item.approval_date))
                ws.cell(row=row, column=13, value=item.requester or "")
                for c in range(1, 15):
                    ws.cell(row=row, column=c).border = _BORDER_THIN
                row += 1

            row += 1  # blank row between work areas

        # Auto-width columns
        for col_idx in range(1, 15):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.column_dimensions["D"].width = 35  # Description wider
        ws.column_dimensions["J"].width = 25  # Assumptions
        ws.column_dimensions["N"].width = 20  # Requester

    buf = io.BytesIO()
    wb.save(buf)
    filename = _build_export_filename(facility_name, "Standard")
    return buf.getvalue(), filename


# ---------------------------------------------------------------------------
# 2. Finance Export (BudgetTemplate format — pixel-perfect)
# ---------------------------------------------------------------------------

# Finance-specific style constants
_FIN_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
_FIN_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="1F3864")
_FIN_HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="000000")
_FIN_HEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
_FIN_MONTH_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_FIN_QTR_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
_FIN_YEAR_FILL = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
_FIN_TOTAL_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FIN_TOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_FIN_DEPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FIN_DEPT_FONT = Font(name="Calibri", bold=True, size=10, color="375623")
_FIN_NUM_FMT = '#,##0.00 "EUR"'
_FIN_DATA_FONT = Font(name="Calibri", size=10)
_FIN_BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
_FIN_BORDER_BOTTOM_MED = Border(
    bottom=Side(style="medium", color="4472C4"),
)
_FIN_ROW1_FONT = Font(name="Calibri", size=9, color="808080")
_FIN_ROW2_FONT = Font(name="Calibri", size=9, color="808080")
_FIN_ROW3_FONT = Font(name="Calibri", bold=True, size=9, color="2F5496")


def _generate_month_range(start_year: int = 2026, start_month: int = 1,
                          end_year: int = 2027, end_month: int = 12) -> list[date]:
    """Generate list of first-of-month dates for the range."""
    months: list[date] = []
    current = date(start_year, start_month, 1)
    end = date(end_year, end_month, 1)
    while current <= end:
        months.append(current)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return months


def _quarter_label(d: date) -> str:
    q = (d.month - 1) // 3 + 1
    return f"Q{q} {d.year}"


def _year_label(d: date) -> str:
    return str(d.year)


async def generate_finance_export(
    session: AsyncSession,
    facility_id: UUID,
    budget_factor: Decimal = DEFAULT_BUDGET_FACTOR,
    start_year: int = 2026,
    start_month: int = 1,
    end_year: int = 2027,
    end_month: int = 12,
) -> tuple[bytes, str]:
    """Generate the pixel-perfect BudgetTemplate-style finance export.

    Row layout matches the original Excel template exactly:
      Row 1: Year numbers (YEAR(date)) above each month column
      Row 2: Month numbers (MONTH(date)) above each month column
      Row 3: Quarter labels ("Q1 2026", ...) above quarter columns
      Row 4: (empty separator)
      Row 5: "TYTAN Technologies" merged title
      Row 6: "CAPITAL EXPENSES" label + monthly/quarterly/yearly sub-headers
      Row 7: Column headers (Identifier, Account, ...)
      Row 8+: Data rows grouped by functional area
      Last:  SUM totals row

    Core formula: IF(item.cash_out == month, amount * factor, 0)

    Returns a tuple of (xlsx_bytes, suggested_filename).
    """
    facility = await _load_facility(session, facility_id)
    facility_name = facility.name if facility else "Unknown"
    functional_areas = await _load_functional_areas(session, facility_id)

    months = _generate_month_range(start_year, start_month, end_year, end_month)
    month_keys = [_month_key(m) for m in months]

    # Build quarter / year indices
    quarters_seen: dict[str, list[int]] = {}
    years_seen: dict[str, list[int]] = {}
    for idx, m in enumerate(months):
        quarters_seen.setdefault(_quarter_label(m), []).append(idx)
        years_seen.setdefault(_year_label(m), []).append(idx)

    quarter_labels = list(dict.fromkeys(_quarter_label(m) for m in months))
    year_labels = list(dict.fromkeys(_year_label(m) for m in months))

    # -- Column layout -------------------------------------------------
    # A-N = 14 metadata columns  (cols 1-14)
    meta_headers = [
        "Identifier", "Account", "Account name", "Category",
        "CC", "CC name", "Ccar", "Ccar name",
        "category", "Description", "Date", "Unit Cost",
        "Useful Life", "Until",
    ]
    NUM_META = len(meta_headers)  # 14

    month_col_start = NUM_META + 1                          # col 15
    qtr_col_start = month_col_start + len(months)           # after months
    yr_col_start = qtr_col_start + len(quarter_labels)      # after quarters
    total_col_end = yr_col_start + len(year_labels) - 1     # last column

    wb = Workbook()
    ws = wb.active
    ws.title = "BudgetTemplate"

    # ================================================================
    # ROW 1 — Year numbers above month columns
    # ================================================================
    for idx, m in enumerate(months):
        col = month_col_start + idx
        cell = ws.cell(row=1, column=col, value=m.year)
        cell.font = _FIN_ROW1_FONT
        cell.alignment = Alignment(horizontal="center")

    # ================================================================
    # ROW 2 — Month numbers above month columns
    # ================================================================
    for idx, m in enumerate(months):
        col = month_col_start + idx
        cell = ws.cell(row=2, column=col, value=m.month)
        cell.font = _FIN_ROW2_FONT
        cell.alignment = Alignment(horizontal="center")

    # ================================================================
    # ROW 3 — Quarter labels above quarter columns
    # ================================================================
    for idx, ql in enumerate(quarter_labels):
        col = qtr_col_start + idx
        cell = ws.cell(row=3, column=col, value=ql)
        cell.font = _FIN_ROW3_FONT
        cell.alignment = Alignment(horizontal="center")

    # Year labels above year columns (also row 3)
    for idx, yl in enumerate(year_labels):
        col = yr_col_start + idx
        cell = ws.cell(row=3, column=col, value=yl)
        cell.font = _FIN_ROW3_FONT
        cell.alignment = Alignment(horizontal="center")

    # ================================================================
    # ROW 4 — empty separator (kept blank)
    # ================================================================

    # ================================================================
    # ROW 5 — "TYTAN Technologies" merged title
    # ================================================================
    last_col_letter = get_column_letter(total_col_end)
    ws.merge_cells(f"A5:{last_col_letter}5")
    title_cell = ws["A5"]
    title_cell.value = "TYTAN Technologies"
    title_cell.font = _FIN_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ================================================================
    # ROW 6 — "CAPITAL EXPENSES" + section sub-headers
    # ================================================================
    ws.cell(row=6, column=1, value="CAPITAL EXPENSES").font = _FIN_SUBTITLE_FONT

    # Sub-header over monthly columns
    if len(months) > 1:
        m_start_letter = get_column_letter(month_col_start)
        m_end_letter = get_column_letter(month_col_start + len(months) - 1)
        ws.merge_cells(f"{m_start_letter}6:{m_end_letter}6")
    m_label_cell = ws.cell(row=6, column=month_col_start, value="Monthly Cash-Out")
    m_label_cell.font = Font(name="Calibri", bold=True, size=9, color="2F5496")
    m_label_cell.alignment = Alignment(horizontal="center")

    # Sub-header over quarter columns
    if len(quarter_labels) > 1:
        q_start_letter = get_column_letter(qtr_col_start)
        q_end_letter = get_column_letter(qtr_col_start + len(quarter_labels) - 1)
        ws.merge_cells(f"{q_start_letter}6:{q_end_letter}6")
    q_label_cell = ws.cell(row=6, column=qtr_col_start, value="Quarterly Totals")
    q_label_cell.font = Font(name="Calibri", bold=True, size=9, color="2F5496")
    q_label_cell.alignment = Alignment(horizontal="center")

    # Sub-header over year columns
    if len(year_labels) > 1:
        y_start_letter = get_column_letter(yr_col_start)
        y_end_letter = get_column_letter(yr_col_start + len(year_labels) - 1)
        ws.merge_cells(f"{y_start_letter}6:{y_end_letter}6")
    y_label_cell = ws.cell(row=6, column=yr_col_start, value="Yearly Totals")
    y_label_cell.font = Font(name="Calibri", bold=True, size=9, color="2F5496")
    y_label_cell.alignment = Alignment(horizontal="center")

    # ================================================================
    # ROW 7 — Column headers
    # ================================================================
    for col_idx, header in enumerate(meta_headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.font = _FIN_HEADER_FONT
        cell.fill = _FIN_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _FIN_BORDER_THIN

    # Month column headers (e.g. "Jan 2026")
    for idx, m in enumerate(months):
        col = month_col_start + idx
        cell = ws.cell(row=7, column=col, value=m.strftime("%b %Y"))
        cell.font = _FIN_HEADER_FONT
        cell.fill = _FIN_MONTH_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _FIN_BORDER_THIN

    # Quarter column headers
    for idx, ql in enumerate(quarter_labels):
        col = qtr_col_start + idx
        cell = ws.cell(row=7, column=col, value=ql)
        cell.font = _FIN_HEADER_FONT
        cell.fill = _FIN_QTR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _FIN_BORDER_THIN

    # Year column headers
    for idx, yl in enumerate(year_labels):
        col = yr_col_start + idx
        cell = ws.cell(row=7, column=col, value=yl)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = _FIN_YEAR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _FIN_BORDER_THIN

    # ================================================================
    # ROW 8+ — Data rows, grouped by functional area
    # ================================================================
    DATA_START_ROW = 8
    row_num = DATA_START_ROW
    item_counter = 1
    fa_group_starts: list[tuple[int, int, str]] = []  # (start_row, end_row, name)

    for fa in functional_areas:
        fa_start = row_num

        # -- FunctionalArea separator row --
        for c in range(1, total_col_end + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = _FIN_DEPT_FILL
            cell.border = _FIN_BORDER_THIN
        ws.cell(row=row_num, column=1, value=fa.name).font = _FIN_DEPT_FONT
        row_num += 1

        fa_has_items = False

        for wa in fa.work_areas:
            for item in wa.cost_items:
                if item.approval_status in EXCLUDED_STATUSES:
                    continue
                fa_has_items = True

                # --- Meta columns ---
                ws.cell(row=row_num, column=1, value=f"CAPEX-{item_counter:04d}")
                ws.cell(row=row_num, column=2, value="")           # Account
                ws.cell(row=row_num, column=3, value="")           # Account name
                ws.cell(row=row_num, column=4,
                        value=_enum_display(item.cost_driver))     # Category
                ws.cell(row=row_num, column=5, value="")           # CC
                ws.cell(row=row_num, column=6, value="")           # CC name
                ws.cell(row=row_num, column=7, value="")           # Ccar
                ws.cell(row=row_num, column=8, value="")           # Ccar name
                ws.cell(row=row_num, column=9,
                        value=_enum_display(item.cost_basis))      # category (cost basis)
                ws.cell(row=row_num, column=10, value=item.description)  # Description
                ws.cell(row=row_num, column=11,
                        value=_date_str(item.expected_cash_out))   # Date

                unit_cost = float(item.total_amount * budget_factor)
                uc_cell = ws.cell(row=row_num, column=12, value=unit_cost)
                uc_cell.number_format = _FIN_NUM_FMT

                ws.cell(row=row_num, column=13, value="")          # Useful Life
                ws.cell(row=row_num, column=14, value="")          # Until

                # Apply base styling to meta cols
                for c in range(1, NUM_META + 1):
                    cell = ws.cell(row=row_num, column=c)
                    cell.font = _FIN_DATA_FONT
                    cell.border = _FIN_BORDER_THIN

                # --- Monthly columns: IF(cash_out == month, amount*factor, 0) ---
                item_month = _month_key(item.expected_cash_out) if item.expected_cash_out else None
                monthly_values: list[float] = []
                for mi, mk in enumerate(month_keys):
                    val = unit_cost if item_month == mk else 0.0
                    monthly_values.append(val)
                    col = month_col_start + mi
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.number_format = _FIN_NUM_FMT
                    cell.font = _FIN_DATA_FONT
                    cell.border = _FIN_BORDER_THIN

                # --- Quarterly aggregation: SUM of 3 months ---
                for q_idx, ql in enumerate(quarter_labels):
                    q_month_indices = quarters_seen[ql]
                    # Build SUM formula referencing the month columns
                    parts = []
                    for mi in q_month_indices:
                        parts.append(f"{get_column_letter(month_col_start + mi)}{row_num}")
                    formula = f"={'+'.join(parts)}"
                    col = qtr_col_start + q_idx
                    cell = ws.cell(row=row_num, column=col, value=formula)
                    cell.number_format = _FIN_NUM_FMT
                    cell.font = _FIN_DATA_FONT
                    cell.border = _FIN_BORDER_THIN

                # --- Yearly totals: SUM over quarter columns ---
                for y_idx, yl in enumerate(year_labels):
                    # Find which quarter columns belong to this year
                    yr_qtr_cols: list[str] = []
                    for q_idx, ql in enumerate(quarter_labels):
                        if ql.endswith(yl):
                            yr_qtr_cols.append(
                                f"{get_column_letter(qtr_col_start + q_idx)}{row_num}"
                            )
                    col = yr_col_start + y_idx
                    if yr_qtr_cols:
                        formula = f"={'+'.join(yr_qtr_cols)}"
                        cell = ws.cell(row=row_num, column=col, value=formula)
                    else:
                        cell = ws.cell(row=row_num, column=col, value=0)
                    cell.number_format = _FIN_NUM_FMT
                    cell.font = _FIN_DATA_FONT
                    cell.border = _FIN_BORDER_THIN

                row_num += 1
                item_counter += 1

        fa_end = row_num - 1 if fa_has_items else fa_start
        fa_group_starts.append((fa_start + 1, fa_end, fa.name))

    # ================================================================
    # TOTALS ROW — SUM formulas over all data rows
    # ================================================================
    total_row = row_num
    first_data = DATA_START_ROW
    last_data = row_num - 1

    # Style the whole totals row
    for c in range(1, total_col_end + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = _FIN_TOTAL_FILL
        cell.font = _FIN_TOTAL_FONT
        cell.border = Border(
            top=Side(style="medium", color="2F5496"),
            bottom=Side(style="medium", color="2F5496"),
        )

    ws.cell(row=total_row, column=1, value="TOTAL")

    if last_data >= first_data:
        # Unit Cost total (col 12)
        letter = get_column_letter(12)
        ws.cell(row=total_row, column=12,
                value=f"=SUM({letter}{first_data}:{letter}{last_data})")
        ws.cell(row=total_row, column=12).number_format = _FIN_NUM_FMT
        ws.cell(row=total_row, column=12).font = _FIN_TOTAL_FONT
        ws.cell(row=total_row, column=12).fill = _FIN_TOTAL_FILL

        # Monthly, quarterly, yearly SUM totals
        for col in range(month_col_start, total_col_end + 1):
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f"=SUM({letter}{first_data}:{letter}{last_data})")
            cell.number_format = _FIN_NUM_FMT
            cell.font = _FIN_TOTAL_FONT
            cell.fill = _FIN_TOTAL_FILL

    # ================================================================
    # Row grouping (outline) by functional area
    # ================================================================
    for grp_start, grp_end, _name in fa_group_starts:
        if grp_end >= grp_start:
            ws.row_dimensions.group(grp_start, grp_end, outline_level=1, hidden=False)

    # ================================================================
    # Column widths
    # ================================================================
    col_widths = {
        1: 14,   # Identifier
        2: 10,   # Account
        3: 16,   # Account name
        4: 14,   # Category
        5: 8,    # CC
        6: 16,   # CC name
        7: 10,   # Ccar
        8: 16,   # Ccar name
        9: 16,   # category
        10: 35,  # Description
        11: 12,  # Date
        12: 16,  # Unit Cost
        13: 12,  # Useful Life
        14: 12,  # Until
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Month columns
    for idx in range(len(months)):
        ws.column_dimensions[get_column_letter(month_col_start + idx)].width = 14

    # Quarter columns
    for idx in range(len(quarter_labels)):
        ws.column_dimensions[get_column_letter(qtr_col_start + idx)].width = 14

    # Year columns
    for idx in range(len(year_labels)):
        ws.column_dimensions[get_column_letter(yr_col_start + idx)].width = 16

    # ================================================================
    # Freeze panes — fix header rows + meta columns
    # ================================================================
    # Freeze at row 8 (first data row) and column O (first month column)
    freeze_cell = f"{get_column_letter(month_col_start)}{DATA_START_ROW}"
    ws.freeze_panes = freeze_cell

    # ================================================================
    # Print setup
    # ================================================================
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3

    buf = io.BytesIO()
    wb.save(buf)
    filename = _build_export_filename(facility_name, "Finance")
    return buf.getvalue(), filename


# ---------------------------------------------------------------------------
# 3. Steering Committee Export
# ---------------------------------------------------------------------------


async def generate_steering_committee_export(
    session: AsyncSession,
    facility_id: UUID,
) -> tuple[bytes, str]:
    """Generate a one-page steering committee summary Excel.

    Contents:
    - Budget vs. Committed vs. Remaining per FunctionalArea
    - Top 10 largest open items
    - Top 5 risk items (no offer / cost_estimation + high amount)
    - Cash-out next 3 months

    Returns a tuple of (xlsx_bytes, suggested_filename).
    """
    facility = await _load_facility(session, facility_id)
    facility_name = facility.name if facility else "Unknown"
    functional_areas = await _load_functional_areas(session, facility_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Steering Committee"

    # --- Title ---
    ws.merge_cells("A1:H1")
    ws["A1"] = "CAPEX Budget — Steering Committee Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F3864")

    today = date.today()
    ws["A2"] = f"Generated: {today.strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(italic=True, color="808080")

    # ======================================================================
    # Section 1: FunctionalArea Overview
    # ======================================================================
    section_row = 4
    ws.cell(row=section_row, column=1, value="Functional Area Overview").font = Font(bold=True, size=13, color="2F5496")

    headers_fa = ["Functional Area", "Budget", "Committed", "Forecast", "Remaining", "% Used"]
    for col_idx, h in enumerate(headers_fa, 1):
        cell = ws.cell(row=section_row + 1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    all_items: list[tuple[CostItem, str]] = []  # (item, fa_name)
    row = section_row + 2
    total_budget = Decimal(0)
    total_committed = Decimal(0)
    total_forecast = Decimal(0)

    for fa in functional_areas:
        budget = _fa_effective_budget(fa)
        committed = sum(
            item.total_amount
            for wa in fa.work_areas
            for item in wa.cost_items
            if item.approval_status == ApprovalStatus.APPROVED
        )
        forecast = sum(
            item.total_amount
            for wa in fa.work_areas
            for item in wa.cost_items
            if item.approval_status not in EXCLUDED_STATUSES
        )
        remaining = budget - forecast
        pct = float(forecast / budget * 100) if budget else 0

        ws.cell(row=row, column=1, value=fa.name)
        ws.cell(row=row, column=2, value=float(budget)).number_format = _NUM_FMT
        ws.cell(row=row, column=3, value=float(committed)).number_format = _NUM_FMT
        ws.cell(row=row, column=4, value=float(forecast)).number_format = _NUM_FMT
        ws.cell(row=row, column=5, value=float(remaining)).number_format = _NUM_FMT
        pct_cell = ws.cell(row=row, column=6, value=pct / 100)
        pct_cell.number_format = '0.0%'

        total_budget += budget
        total_committed += committed
        total_forecast += forecast

        for wa in fa.work_areas:
            for item in wa.cost_items:
                all_items.append((item, fa.name))

        row += 1

    # Total row
    total_remaining = total_budget - total_forecast
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(total_budget)).number_format = _NUM_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_committed)).number_format = _NUM_FMT
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_forecast)).number_format = _NUM_FMT
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_remaining)).number_format = _NUM_FMT
    ws.cell(row=row, column=5).font = Font(bold=True)
    if total_budget:
        pct_cell = ws.cell(row=row, column=6, value=float(total_forecast / total_budget))
        pct_cell.number_format = '0.0%'
        pct_cell.font = Font(bold=True)

    row += 2

    # ======================================================================
    # Section 2: Top 10 Largest Open Items
    # ======================================================================
    ws.cell(row=row, column=1, value="Top 10 Largest Open Items").font = Font(bold=True, size=13, color="2F5496")
    row += 1

    top10_headers = ["#", "Functional Area", "Description", "Amount", "Status", "Cash Out"]
    for col_idx, h in enumerate(top10_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row += 1

    open_statuses = {ApprovalStatus.OPEN, ApprovalStatus.SUBMITTED_FOR_APPROVAL,
                     ApprovalStatus.ON_HOLD, ApprovalStatus.PENDING_SUPPLIER_NEGOTIATION,
                     ApprovalStatus.PENDING_TECHNICAL_CLARIFICATION}
    open_items = [(i, dn) for i, dn in all_items if i.approval_status in open_statuses]
    open_items.sort(key=lambda x: x[0].total_amount, reverse=True)

    for rank, (item, fa_name) in enumerate(open_items[:10], 1):
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=fa_name)
        ws.cell(row=row, column=3, value=item.description)
        ws.cell(row=row, column=4, value=float(item.total_amount)).number_format = _NUM_FMT
        ws.cell(row=row, column=5, value=_enum_display(item.approval_status))
        ws.cell(row=row, column=6, value=_date_str(item.expected_cash_out))
        row += 1

    row += 1

    # ======================================================================
    # Section 3: Top 5 Risk Items
    # ======================================================================
    ws.cell(row=row, column=1, value="Top 5 Risk Items").font = Font(bold=True, size=13, color="2F5496")
    ws.cell(row=row, column=3, value="(No supplier offer + high amount)").font = Font(italic=True, color="808080", size=9)
    row += 1

    risk_headers = ["#", "Functional Area", "Description", "Amount", "Cost Basis", "Status"]
    for col_idx, h in enumerate(risk_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    row += 1

    # Risk = cost_estimation basis (no supplier offer) + in open statuses
    risk_items = [
        (i, dn) for i, dn in all_items
        if i.cost_basis == "cost_estimation" and i.approval_status in open_statuses
    ]
    risk_items.sort(key=lambda x: x[0].total_amount, reverse=True)

    for rank, (item, fa_name) in enumerate(risk_items[:5], 1):
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=fa_name)
        ws.cell(row=row, column=3, value=item.description)
        ws.cell(row=row, column=4, value=float(item.total_amount)).number_format = _NUM_FMT
        ws.cell(row=row, column=5, value=_enum_display(item.cost_basis))
        ws.cell(row=row, column=6, value=_enum_display(item.approval_status))
        row += 1

    row += 1

    # ======================================================================
    # Section 4: Cash-Out Next 3 Months
    # ======================================================================
    ws.cell(row=row, column=1, value="Cash-Out — Next 3 Months").font = Font(bold=True, size=13, color="2F5496")
    row += 1

    # Determine next 3 months
    next_months: list[date] = []
    current_month = date(today.year, today.month, 1)
    for _ in range(3):
        next_months.append(current_month)
        if current_month.month == 12:
            current_month = date(current_month.year + 1, 1, 1)
        else:
            current_month = date(current_month.year, current_month.month + 1, 1)

    cashout_headers = ["Functional Area"] + [m.strftime("%b %Y") for m in next_months] + ["Total"]
    for col_idx, h in enumerate(cashout_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row += 1

    month_keys_next = [_month_key(m) for m in next_months]
    grand_totals = [Decimal(0)] * len(next_months)

    for fa in functional_areas:
        ws.cell(row=row, column=1, value=fa.name)
        fa_month_totals = [Decimal(0)] * len(next_months)

        for wa in fa.work_areas:
            for item in wa.cost_items:
                if item.approval_status in EXCLUDED_STATUSES:
                    continue
                if item.expected_cash_out:
                    item_mk = _month_key(item.expected_cash_out)
                    for mi, mk in enumerate(month_keys_next):
                        if item_mk == mk:
                            fa_month_totals[mi] += item.total_amount

        fa_total = Decimal(0)
        for mi, val in enumerate(fa_month_totals):
            ws.cell(row=row, column=2 + mi, value=float(val)).number_format = _NUM_FMT
            grand_totals[mi] += val
            fa_total += val

        ws.cell(row=row, column=2 + len(next_months), value=float(fa_total)).number_format = _NUM_FMT
        row += 1

    # Grand total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    grand_sum = Decimal(0)
    for mi, val in enumerate(grand_totals):
        ws.cell(row=row, column=2 + mi, value=float(val)).number_format = _NUM_FMT
        ws.cell(row=row, column=2 + mi).font = Font(bold=True)
        grand_sum += val
    ws.cell(row=row, column=2 + len(next_months), value=float(grand_sum)).number_format = _NUM_FMT
    ws.cell(row=row, column=2 + len(next_months)).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    for c in range(4, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    filename = _build_export_filename(facility_name, "Steering_Committee")
    return buf.getvalue(), filename
